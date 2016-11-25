import requests
from lxml import etree
import io
from bs4 import BeautifulSoup
import json
import random
from openpyxl import Workbook
import argparse


def get_courses_list():
    courses_xml = requests.get('https://www.coursera.org/sitemap~www~courses.xml').content
    tree = etree.parse(io.BytesIO(courses_xml))
    root = tree.getroot()
    courses = []
    for element in root.iter():
        for subel in element:
            if 'https://' in subel.text and len(subel.text) > 0:
                courses.append(subel.text)
    return courses


def get_soup_for_course(course_url):
    course_html = requests.get(course_url).text
    soup = BeautifulSoup(course_html, "html.parser")
    return soup


def get_course_rating(course_soup):
    rating = None
    divs = course_soup.find_all('div')
    for div in divs:
        div_classes = div.get('class')
        if div_classes is not None:
            if 'ratings-text' in div_classes and 'bt3-visible-xs' in div_classes:
                rating = div.string
    return rating


def get_course_language(course_soup):
    language = None
    div_lang = course_soup.find('div', {'class': 'language-info'})
    if div_lang is not None:
        language = div_lang.text
    return language


def get_course_name(course_soup):
    name = None
    script = course_soup.find('script', {'type': 'application/ld+json'})
    if script and 'name' in script.text:
        name = json.loads(script.text)['name']
    return name


def get_course_start_date(course_soup):
    start_date = None
    script = course_soup.find('script', {'type': 'application/ld+json'})
    if script and 'startDate' in script.text:
        start_date = json.loads(script.text)['hasCourseInstance'][0]['startDate']
    return start_date


def get_course_length(course_soup):
    week_divs = course_soup.find_all('div', {'class': 'week'})
    return len(week_divs)


def get_course_info(course_soup):
    lang = get_course_language(course_soup)
    rating = get_course_rating(course_soup)
    start_date = get_course_start_date(course_soup)
    name = get_course_name(course_soup)
    length = get_course_length(course_soup)
    return {'name': name, 'lang': lang, 'rating': rating, 'start': start_date, 'len': length}


def output_courses_info_to_xlsx(filepath, courses_info):
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Название курса'
    ws['B1'] = 'Язык курса'
    ws['C1'] = 'Дата начала'
    ws['D1'] = 'Продолжительность'
    ws['E1'] = 'Рейтинг'
    for course in courses_info:
        ws.append([course['name'], course['lang'], course['start'], course['len'], course['rating']])
    wb.save(filepath)


def get_filepath():
    parser = argparse.ArgumentParser()
    parser.add_argument('-f', '--filepath',
                        required=False,
                        help='Path to the output file')
    args = parser.parse_args()
    if not args.filepath:
        file_path = input('Введите имя файла для таблицы с курсами -->')
    else:
        file_path = args.filepath
    return file_path

if __name__ == '__main__':

    course_list = get_courses_list()
    courses_info = []
    for i in range(19):
        random_course_url = course_list[random.randrange(0, len(course_list)-1)]
        course_soup = get_soup_for_course(random_course_url)
        courses_info.append(get_course_info(course_soup))
    filepath = get_filepath()
    output_courses_info_to_xlsx(filepath, courses_info)


